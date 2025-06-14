import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, unquote, parse_qs, urlparse
import os
import time
import re
import csv
import mimetypes
import configparser # 설정 파일 처리
import sys # 폴더 열기용 (macOS/Linux 구분)
from eden_data_preprocess_gui import make_roulette_csv_from_excel

# --- 기본 설정 ---
BASE_URL = "https://anothereden.wiki"
TARGET_URL = "https://anothereden.wiki/w/Characters"
CONFIG_FILE = "scraper_config.ini"

# 전역 변수로 출력 경로 관련 변수들을 초기화합니다.
# 이 값들은 load_config 및 GUI를 통해 설정됩니다.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_BASE_DIR = SCRIPT_DIR # 기본값은 스크립트 위치
IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")
EXCEL_FILENAME_BASE = "another_eden_characters_detailed" # 확장자 제외 기본 이름

# --- 설정 파일 처리 ---
def load_config():
    global OUTPUT_BASE_DIR, IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV
    config = configparser.ConfigParser()
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    if os.path.exists(config_path):
        config.read(config_path)
        OUTPUT_BASE_DIR = config.get('Paths', 'OutputDirectory', fallback=SCRIPT_DIR)
    else:
        OUTPUT_BASE_DIR = SCRIPT_DIR # 설정 파일 없으면 스크립트 위치
    
    # 경로 변수들 업데이트
    IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")
    # EXCEL_FILENAME은 저장 시점에 최종 결정됨

def save_config():
    global OUTPUT_BASE_DIR
    config = configparser.ConfigParser()
    config['Paths'] = {'OutputDirectory': OUTPUT_BASE_DIR}
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    with open(config_path, 'w') as configfile:
        config.write(configfile)

# --- 이미지 및 폴더 처리 함수 ---
def setup_directories():
    # IMAGE_DIR은 OUTPUT_BASE_DIR 기준으로 설정됨
    if not os.path.exists(IMAGE_DIR):
        os.makedirs(IMAGE_DIR)
        # log_queue.put(f"Created directory: {IMAGE_DIR}") # GUI 앱에서는 log_queue 사용
    icons_path = os.path.join(IMAGE_DIR, "icons")
    elements_path = os.path.join(IMAGE_DIR, "elements_equipment")
    if not os.path.exists(icons_path):
        os.makedirs(icons_path)
    if not os.path.exists(elements_path):
        os.makedirs(elements_path)

def get_unique_filename(filepath):
    """파일 경로가 중복될 경우 (숫자)를 붙여 고유한 경로 반환"""
    if not os.path.exists(filepath):
        return filepath
    
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath):
            return new_filepath
        counter += 1

def download_image(image_url, subfolder=""):
    # (이전 GUI 코드의 download_image 함수와 거의 동일)
    if not image_url:
        return None
    full_image_url = urljoin(BASE_URL, image_url)
    try:
        parsed_url = urlparse(full_image_url)
        query_params = parse_qs(parsed_url.query)
        image_name_from_f = query_params.get('f', [None])[0]
        
        if image_name_from_f:
            image_name = os.path.basename(unquote(image_name_from_f))
        else:
            image_name = os.path.basename(unquote(parsed_url.path.split('?')[0]))

        if not image_name or image_name.lower() == "thumb.php" or image_name.lower() == "index.php":
            temp_name = unquote(full_image_url.split('/')[-1].split('?')[0])
            image_name = (temp_name[:50] + ".png") if temp_name else "unknown_image.png"

        base_name, ext = os.path.splitext(image_name)
        if not ext or len(ext) > 5: # 확장자가 너무 길거나 없는 경우
            try:
                head_resp = requests.head(full_image_url, timeout=3, allow_redirects=True)
                head_resp.raise_for_status()
                content_type = head_resp.headers.get('Content-Type')
                if content_type:
                    guessed_ext = mimetypes.guess_extension(content_type.split(';')[0])
                    image_name = base_name + (guessed_ext if guessed_ext and guessed_ext != '.jpe' else ('.jpg' if guessed_ext == '.jpe' else '.png'))
                else: image_name = base_name + ".png" 
            except: image_name = base_name + ".png"
        
        image_name = re.sub(r'[<>:"/\\|?*]', '_', image_name)
        image_name = image_name[:200] # 파일 이름 길이 제한

        save_path_dir = os.path.join(IMAGE_DIR, subfolder) # IMAGE_DIR은 OUTPUT_BASE_DIR 기준으로 설정됨
        save_path = os.path.join(save_path_dir, image_name)
        
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return save_path

        img_response = requests.get(full_image_url, stream=True, timeout=10)
        img_response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in img_response.iter_content(8192):
                f.write(chunk)
        time.sleep(0.05) 
        return save_path
    except requests.exceptions.RequestException as e:
        # log_queue에 메시지 전송
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue:
            error_log_queue.put(f"Download Error (Network) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    except Exception as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue:
            error_log_queue.put(f"Download Error (Other) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    return None

# --- 스크레이핑 핵심 로직 ---
def scraping_logic(log_queue_ref, progress_queue_ref, generate_structure_sheet_only_mode, selected_output_dir):
    # 스레드에서 log_queue와 progress_queue에 접근할 수 있도록 참조 전달
    threading.current_thread().log_queue_ref = log_queue_ref
    threading.current_thread().progress_queue_ref = progress_queue_ref
    
    # 전역 경로 변수들을 이 스레드 내의 로컬 변수로 업데이트 (중요)
    global IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV 
    IMAGE_DIR = os.path.join(selected_output_dir, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(selected_output_dir, "structure_analysis.csv")
    
    setup_directories()
    log_queue_ref.put(f"Output directory set to: {selected_output_dir}")
    log_queue_ref.put(f"Fetching page: {TARGET_URL}")
    
    headers_ua = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    try:
        response = requests.get(TARGET_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        log_queue_ref.put(f"Error fetching page: {e}")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Page fetched. Parsing HTML...")
    soup = BeautifulSoup(response.content, 'html.parser')
    
    char_table = soup.find('table', class_='chara-table')
    if not char_table:
        char_table = soup.find('table', class_='wikitable') 
        if not char_table:
            log_queue_ref.put("Character table not found. Please verify table class name.")
            progress_queue_ref.put({'done': True, 'error': True})
            return

    log_queue_ref.put("Character table found. Parsing rows...")
    
    all_character_data_for_final_excel = []
    raw_data_for_structure_analysis = []
    
    rows = char_table.find_all('tr')
    total_rows_to_process = len(rows) -1 
    log_queue_ref.put(f"Found {total_rows_to_process} potential character rows.")
    if total_rows_to_process <= 0 :
        log_queue_ref.put("No character rows found to process.")
        progress_queue_ref.put({'done': True})
        return
        
    progress_queue_ref.put({'max': total_rows_to_process, 'value': 0})

    for i, row in enumerate(rows[1:]): 
        current_progress = i + 1
        progress_queue_ref.put({'value': current_progress}) 

        cells = row.find_all('td')
        if len(cells) < 4:
            continue

        analysis_row_data = {'row_index': current_progress}
        try:
            icon_img_tag_analysis = cells[0].find('img')
            analysis_row_data['cell_1_icon_src'] = icon_img_tag_analysis['src'] if icon_img_tag_analysis and icon_img_tag_analysis.get('src') else "N/A"
            analysis_row_data['cell_1_icon_alt'] = icon_img_tag_analysis['alt'] if icon_img_tag_analysis and icon_img_tag_analysis.get('alt') else "N/A"
            # HTML은 너무 길어질 수 있으므로, 필요한 정보만 추출
            # analysis_row_data['cell_1_icon_html'] = cells[0].prettify()[:300] 
            
            name_tag_analysis = cells[1].find('a')
            analysis_row_data['cell_2_name_link_text'] = name_tag_analysis.text.strip() if name_tag_analysis else "N/A"
            analysis_row_data['cell_2_full_text'] = cells[1].get_text(separator=' | ').strip()[:300]
            # analysis_row_data['cell_2_name_rarity_html'] = cells[1].prettify()[:300]

            ee_imgs_analysis = cells[2].find_all('img')
            analysis_row_data['cell_3_img_srcs'] = ", ".join([img['src'] for img in ee_imgs_analysis if img.get('src')])[:300]
            analysis_row_data['cell_3_img_alts'] = ", ".join([img.get('alt', '') for img in ee_imgs_analysis])[:300]
            # analysis_row_data['cell_3_elem_equip_html'] = cells[2].prettify()[:300]

            analysis_row_data['cell_4_text'] = cells[3].text.strip()[:300]
            # analysis_row_data['cell_4_release_date_html'] = cells[3].prettify()[:300]
            raw_data_for_structure_analysis.append(analysis_row_data)
        except Exception as e:
            log_queue_ref.put(f"Row {current_progress}: Error preparing structure analysis data: {e}")
            raw_data_for_structure_analysis.append({'row_index': current_progress, 'error_collecting_data': str(e)})

        if generate_structure_sheet_only_mode:
            if current_progress % 50 == 0: log_queue_ref.put(f"Structure Analysis: Processed row {current_progress}/{total_rows_to_process}")
            continue

        try:
            icon_cell = cells[0]
            icon_img_tag = icon_cell.find('img')
            icon_src = icon_img_tag['src'] if icon_img_tag and icon_img_tag.get('src') else None
            icon_alt = icon_img_tag['alt'] if icon_img_tag and icon_img_tag.get('alt') else ""
            icon_local_path = download_image(icon_src, "icons") if icon_src else None
            # [추가] 아이콘 파일명 추출 (URL에서 파일명만 추출)
            icon_filename = None
            if icon_src:
                parsed_url = urlparse(icon_src)
                query_params = parse_qs(parsed_url.query)
                image_name_from_f = query_params.get('f', [None])[0]
                if image_name_from_f:
                    icon_filename = os.path.basename(unquote(image_name_from_f)).replace(' ', '_')
                else:
                    icon_filename = os.path.basename(unquote(parsed_url.path.split('?')[0])).replace(' ', '_')
            name_rarity_cell = cells[1]
            name_tag = name_rarity_cell.find('a')
            name = name_tag.text.strip() if name_tag else ""
            rarity = ""
            lines_in_cell = [line.strip() for line in name_rarity_cell.get_text(separator='\n').splitlines() if line.strip()]
            if not name and lines_in_cell: name = lines_in_cell[0]
            for line_text in reversed(lines_in_cell):
                if "★" in line_text:
                    rarity = line_text
                    break
            if not rarity:
                full_text_for_rarity = name_rarity_cell.get_text(separator=" ").strip()
                rarity_match = re.search(r'\d(?:~\d)?★(?:\s*\S+)?', full_text_for_rarity)
                if rarity_match: rarity = rarity_match.group(0).strip()
            if not name and icon_alt and "Icon" in icon_alt: name = icon_alt.replace(" Icon","").strip()
            elif not name and analysis_row_data.get('cell_2_name_link_text', "N/A") != "N/A": name = analysis_row_data['cell_2_name_link_text']
            
            element_equipment_cell = cells[2]
            ee_icon_tags = element_equipment_cell.find_all('img')
            element_equipment_icon_paths = []
            element_equipment_icon_alts = []
            for img_tag in ee_icon_tags:
                ee_src = img_tag.get('src')
                ee_alt = img_tag.get('alt', "") 
                if ee_src:
                    local_path = download_image(ee_src, "elements_equipment")
                    if local_path:
                        element_equipment_icon_paths.append(local_path)
                        element_equipment_icon_alts.append(ee_alt)
            
            release_date_cell = cells[3]
            release_date = release_date_cell.text.strip()

            if name or icon_local_path:
                all_character_data_for_final_excel.append({
                    "icon_path": icon_local_path,
                    "icon_filename": icon_filename,  # [추가] 아이콘 파일명 컬럼
                    "name": name,
                    "rarity": rarity,
                    "element_equipment_paths": element_equipment_icon_paths,
                    "element_equipment_alts": element_equipment_icon_alts,
                    "release_date": release_date
                })
                if current_progress % 20 == 0 : log_queue_ref.put(f"Row {current_progress}: Parsed & Downloaded for '{name}'")
        except Exception as e:
            log_queue_ref.put(f"Row {current_progress}: Error parsing for final Excel: {e}")
            continue
    
    if raw_data_for_structure_analysis:
        log_queue_ref.put(f"Saving structure analysis to: {STRUCTURE_ANALYSIS_FILENAME_CSV}")
        try:
            headers_csv = raw_data_for_structure_analysis[0].keys() if raw_data_for_structure_analysis else []
            if headers_csv:
                unique_csv_path = get_unique_filename(STRUCTURE_ANALYSIS_FILENAME_CSV)
                with open(unique_csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=headers_csv)
                    writer.writeheader()
                    writer.writerows(raw_data_for_structure_analysis)
                log_queue_ref.put(f"Structure analysis sheet saved: {unique_csv_path}")
            else:
                log_queue_ref.put("No data for structure analysis sheet.")
        except Exception as e:
            log_queue_ref.put(f"Error writing CSV file {STRUCTURE_ANALYSIS_FILENAME_CSV}: {e}")

    if generate_structure_sheet_only_mode:
        log_queue_ref.put("Structure analysis sheet generated. Main Excel generation skipped.")
        progress_queue_ref.put({'done': True})
        return
        
    if not all_character_data_for_final_excel:
        log_queue_ref.put("No data to save to the final Excel sheet.")
        progress_queue_ref.put({'done': True})
        return

    final_excel_path_base = os.path.join(selected_output_dir, EXCEL_FILENAME_BASE)
    final_excel_full_path = get_unique_filename(f"{final_excel_path_base}.xlsx")
    
    log_queue_ref.put(f"Saving final data to Excel: {final_excel_full_path} ({len(all_character_data_for_final_excel)} characters)")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"
    max_ee_icons = 0
    if all_character_data_for_final_excel:
        paths_lengths = [len(char_data["element_equipment_paths"]) for char_data in all_character_data_for_final_excel]
        if paths_lengths: max_ee_icons = max(paths_lengths)

    headers_excel = ["Icon", "Icon Filename", "Name", "Rarity"]  # [추가] 헤더에 아이콘 파일명
    for i in range(max_ee_icons):
        headers_excel.extend([f"Elem/Equip {i+1} Icon", f"Elem/Equip {i+1} Alt"])
    headers_excel.append("Release Date")
    ws.append(headers_excel)

    ws.column_dimensions[get_column_letter(1)].width = 12 
    ws.column_dimensions[get_column_letter(2)].width = 35 
    ws.column_dimensions[get_column_letter(3)].width = 15 
    for i in range(max_ee_icons):
        ws.column_dimensions[get_column_letter(4 + i*2)].width = 12 
        ws.column_dimensions[get_column_letter(5 + i*2)].width = 25 
    ws.column_dimensions[get_column_letter(4 + max_ee_icons*2)].width = 15

    for row_idx_excel, char_data in enumerate(all_character_data_for_final_excel, start=2):
        ws.row_dimensions[row_idx_excel].height = 60 
        current_col = 1 
        if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
            try:
                img = OpenpyxlImage(char_data["icon_path"])
                img.height = 75 
                img.width = 75  
                ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
            except Exception as e:
                log_queue_ref.put(f"Excel Write Error (Icon) for {char_data.get('name', 'N/A')}: {e}")
                ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
        current_col += 1
        # [추가] 아이콘 파일명 컬럼
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("icon_filename", ""))
        current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("name", "N/A"))
        current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("rarity", "N/A"))
        current_col += 1

        for i in range(max_ee_icons):
            if i < len(char_data["element_equipment_paths"]):
                icon_path = char_data["element_equipment_paths"][i]
                icon_alt = char_data["element_equipment_alts"][i]
                if icon_path and os.path.exists(icon_path):
                    try:
                        img = OpenpyxlImage(icon_path)
                        img.height = 30 
                        img.width = 30
                        ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                    except Exception as e:
                        log_queue_ref.put(f"Excel Write Error (E/E Icon) for {char_data.get('name', 'N/A')}: {e}")
                        ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
                ws.cell(row=row_idx_excel, column=current_col + 1, value=icon_alt)
            else: 
                ws.cell(row=row_idx_excel, column=current_col, value="")
                ws.cell(row=row_idx_excel, column=current_col + 1, value="")
            current_col += 2 
        
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("release_date", "N/A"))
        if (row_idx_excel-1) % 50 == 0 : log_queue_ref.put(f"Excel: Wrote data for '{char_data.get('name')}' (Processed {row_idx_excel-1} characters)")

    try:
        wb.save(final_excel_full_path)
        log_queue_ref.put(f"Final data successfully saved to {final_excel_full_path}")
        # 자동 전처리: Excel → CSV (eden_roulette_data.csv)
        try:
            csv_generated_path = make_roulette_csv_from_excel(final_excel_full_path)
            log_queue_ref.put(f"Roulette CSV generated: {csv_generated_path}")
        except Exception as e_pre:
            log_queue_ref.put(f"Error generating roulette CSV: {e_pre}")
    except Exception as e:
        log_queue_ref.put(f"Error saving final Excel file: {e}")
        # GUI 스레드에서 messagebox를 호출해야 하므로, 여기서는 로그만 남김
        # 실제로는 progress_queue를 통해 에러 상태를 전달하고 GUI에서 처리
        progress_queue_ref.put({'done': True, 'error': True, 'error_message': f"Excel 저장 오류:\n{e}\n\n파일이 다른 프로그램에서 열려있는지 확인하세요."})
        return # 오류 발생 시 여기서 종료
    
    progress_queue_ref.put({'done': True})

# --- Tkinter GUI Application ---
class ScraperApp:
    def __init__(self, root_window):
        self.root = root_window
        root_window.title("어나더에덴 캐릭터 스크레이퍼")
        root_window.geometry("800x600")

        load_config() # 앱 시작 시 설정 로드
        self.output_dir_var = tk.StringVar(value=OUTPUT_BASE_DIR)


        style = ttk.Style()
        try:
            if os.name == 'nt': style.theme_use('vista')
            else: style.theme_use('clam') 
        except tk.TclError: style.theme_use('clam')

        # 최상위 프레임 (패딩용)
        main_frame = ttk.Frame(root_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 경로 설정 프레임
        path_frame = ttk.LabelFrame(main_frame, text="출력 설정", padding="10")
        path_frame.pack(fill=tk.X, pady=(0,10))

        ttk.Label(path_frame, text="저장 폴더:").grid(row=0, column=0, padx=(0,5), pady=5, sticky="w")
        self.output_dir_entry = ttk.Entry(path_frame, textvariable=self.output_dir_var, width=60)
        self.output_dir_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.btn_browse_dir = ttk.Button(path_frame, text="폴더 찾아보기", command=self.browse_output_directory)
        self.btn_browse_dir.grid(row=0, column=2, padx=(5,0), pady=5)
        path_frame.columnconfigure(1, weight=1) # Entry 위젯이 늘어나도록

        # 컨트롤 프레임
        control_frame = ttk.Frame(main_frame) # 패딩은 개별 버튼/위젯에
        control_frame.pack(fill=tk.X, pady=(0,10))


        self.btn_structure = ttk.Button(control_frame, text="1. 구조 분석 시트 생성", command=lambda: self.start_scraping_thread(True))
        self.btn_structure.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        self.btn_full_report = ttk.Button(control_frame, text="2. 최종 보고서 생성 (이미지 포함)", command=lambda: self.start_scraping_thread(False))
        self.btn_full_report.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        self.btn_open_folder = ttk.Button(control_frame, text="출력 폴더 열기", command=self.open_output_folder)
        self.btn_open_folder.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        # 로그 프레임
        log_label_frame = ttk.LabelFrame(main_frame, text="진행 상황 및 로그", padding="10")
        log_label_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_label_frame, wrap=tk.WORD, state=tk.DISABLED, height=15, font=("Malgun Gothic", 9) if os.name == 'nt' else ("TkDefaultFont", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.tag_configure("error", foreground="red")
        self.log_text.tag_configure("info", foreground="blue")
        self.log_text.tag_configure("success", foreground="green")

        self.progress_bar = ttk.Progressbar(log_label_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(pady=10, fill=tk.X, padx=5)

        self.log_queue = queue.Queue()
        self.progress_queue = queue.Queue()
        self.root.after(100, self.process_queues)
        
        # 창 닫을 때 설정 저장
        root_window.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        save_config() # 현재 경로 저장
        self.root.destroy()

    def browse_output_directory(self):
        global OUTPUT_BASE_DIR # 전역변수 업데이트
        selected_dir = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if selected_dir:
            self.output_dir_var.set(selected_dir)
            OUTPUT_BASE_DIR = selected_dir # 전역변수에도 반영
            self.log_message(f"출력 폴더가 다음으로 설정되었습니다: {selected_dir}", "info")
            save_config() # 변경 시 바로 저장

    def open_output_folder(self):
        folder_to_open = self.output_dir_var.get()
        if not os.path.isdir(folder_to_open):
            self.log_message(f"폴더를 찾을 수 없습니다: {folder_to_open}", "error")
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{folder_to_open}")
            return
        try:
            if os.name == 'nt': os.startfile(folder_to_open)
            elif sys.platform == "darwin": os.system(f'open "{folder_to_open}"')
            else: os.system(f'xdg-open "{folder_to_open}"')
            self.log_message(f"출력 폴더가 열렸습니다: {folder_to_open}", "info")
        except Exception as e:
            self.log_message(f"폴더 열기 오류: {e}", "error")
            messagebox.showerror("폴더 열기 오류", f"출력 폴더를 여는 중 오류 발생:\n{e}")

    def log_message(self, message, tag=None):
        self.log_text.config(state=tk.NORMAL)
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        if tag: self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        else: self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END) 

    def start_scraping_thread(self, structure_only_mode):
        current_output_dir = self.output_dir_var.get()
        if not os.path.isdir(current_output_dir):
            messagebox.showerror("오류", f"유효한 출력 폴더를 선택해주세요.\n현재 경로: {current_output_dir}")
            return

        self.btn_structure.config(state=tk.DISABLED)
        self.btn_full_report.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        self.log_text.config(state=tk.NORMAL); self.log_text.delete('1.0', tk.END); self.log_text.config(state=tk.DISABLED)
        self.log_message("Scraping process started...", "info")

        save_config() # 작업 시작 전 현재 경로 저장

        self.thread = threading.Thread(target=scraping_logic, 
                                       args=(self.log_queue, self.progress_queue, structure_only_mode, current_output_dir),
                                       daemon=True)
        self.thread.start()

    def process_queues(self):
        try:
            while True: 
                message = self.log_queue.get_nowait()
                tag_to_use = None
                if "Error" in message or "오류" in message : tag_to_use = "error"
                elif "saved" in message or "완료" in message or "successfully" in message : tag_to_use = "success"
                elif "Fetching" in message or "Parsing" in message or "set to" in message : tag_to_use = "info"
                self.log_message(message, tag_to_use)
        except queue.Empty: pass

        try:
            while True: 
                progress_data = self.progress_queue.get_nowait()
                if 'max' in progress_data: self.progress_bar["maximum"] = progress_data['max']
                if 'value' in progress_data: self.progress_bar["value"] = progress_data['value']
                if 'done' in progress_data and progress_data['done']:
                    self.btn_structure.config(state=tk.NORMAL)
                    self.btn_full_report.config(state=tk.NORMAL)
                    if self.progress_bar["value"] < self.progress_bar["maximum"]:
                        self.progress_bar["value"] = self.progress_bar["maximum"] 
                    
                    if 'error' in progress_data and progress_data['error']:
                        msg = progress_data.get('error_message', "작업 중 오류가 발생했습니다. 로그를 확인하세요.")
                        self.log_message(f"Process finished with errors: {msg}", "error")
                        messagebox.showerror("오류", msg)
                    else:
                        self.log_message("Process finished successfully.", "success")
                        messagebox.showinfo("완료", "작업이 성공적으로 완료되었습니다!")
        except queue.Empty: pass
        self.root.after(100, self.process_queues)

if __name__ == "__main__":
    root = tk.Tk()
    app = ScraperApp(root)
    root.mainloop()